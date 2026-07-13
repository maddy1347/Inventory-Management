"""
Inventory Scoring + SOV Suggestion Agent
Tag-Level Pipeline v4 (Dual-Side + Bundle Blocking + Revenue Loss Cap)

Single source of truth for the scoring pipeline AND the colour-coded workbook
builder. Both run_agent.py (CLI) and docs/assets/pipeline.py (browser copy,
run under Pyodide) import this exact file -- keep them in sync (see
scripts/sync_webapp.py, run automatically by CI).

No live capping API is assumed, so "current cap" and cap-increase suggestions
are derived only from the data available in the file (Avg_Daily_Requests *
multiplier). If you later connect a capping API, see the CURRENT_CAP hook
noted near Daily_Cap below.
"""

import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LADDER = [0.00, 0.25, 0.50, 0.75, 1.00, 1.25, 1.50, 1.75]


# --------------------------------------------------------------------------
# Column auto-detection (Section 1.2)
# --------------------------------------------------------------------------
def _normalize(name: str) -> str:
    return re.sub(r'[\s_]+', '', name.strip().lower())


CANDIDATES = {
    'supply_tag': ['supplytag', 'supply'],
    'demand_tag': ['demandtag', 'demand'],
    'bundle': ['appbundle', 'bundle', 'bundleid'],
    'date': ['date'],
    'requests': ['requests', 'request', 'adrequests', 'adrequest'],
    'revenue': ['revenue'],
    'fillrate': ['fillrate', 'fill_rate', 'fill'],
    'impressions': ['impressions', 'impression', 'imps'],
}


def detect_columns(df: pd.DataFrame) -> dict:
    norm_map = {_normalize(c): c for c in df.columns}
    found = {}
    for key, aliases in CANDIDATES.items():
        for alias in aliases:
            if alias in norm_map:
                found[key] = norm_map[alias]
                break
    return found


# --------------------------------------------------------------------------
# Robust stat helper
# --------------------------------------------------------------------------
def robust_std(x: np.ndarray) -> float:
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return 1e-9
    med = np.median(x)
    return max(1.4826 * np.median(np.abs(x - med)), 1e-9)


def pct_rank(series: pd.Series) -> pd.Series:
    """0..1 percentile rank, average method, matches spec's percentile_rank()."""
    if len(series) <= 1:
        return pd.Series(np.full(len(series), 0.5), index=series.index)
    return series.rank(method='average', pct=True)


# --------------------------------------------------------------------------
# Config container
# --------------------------------------------------------------------------
class Config:
    def __init__(
        self,
        from_date=None,
        to_date=None,
        cost_enabled=False,
        cost_basis='requests',       # 'requests' or 'impressions'
        cost_per_1m_req=0.0,
        cost_per_1k_imp=0.0,
        rev_threshold_enabled=False,
        rev_threshold=0.0,
        max_rev_loss_enabled=False,
        max_rev_loss_pct=0.0,
    ):
        self.from_date = from_date
        self.to_date = to_date
        self.cost_enabled = cost_enabled
        self.cost_basis = cost_basis
        self.cost_per_1m_req = cost_per_1m_req
        self.cost_per_1k_imp = cost_per_1k_imp
        self.rev_threshold_enabled = rev_threshold_enabled
        self.rev_threshold = rev_threshold
        self.max_rev_loss_enabled = max_rev_loss_enabled
        self.max_rev_loss_pct = max_rev_loss_pct


# --------------------------------------------------------------------------
# Row level cleaning + derivation (Step 4.3 + row-level part of 4.4A)
# --------------------------------------------------------------------------
def clean_rows(df: pd.DataFrame, cols: dict) -> pd.DataFrame:
    df = df.copy()
    df['Date'] = pd.to_datetime(df[cols['date']], errors='coerce')

    for key in ('requests', 'revenue', 'fillrate', 'impressions'):
        if key in cols:
            df[key] = pd.to_numeric(df[cols[key]], errors='coerce')
        else:
            df[key] = np.nan

    for key in ('requests', 'revenue', 'fillrate', 'impressions'):
        df[key] = df[key].fillna(0.0).clip(lower=0.0)

    if 'impressions' not in cols and 'fillrate' in cols:
        df['impressions'] = df['fillrate'] * df['requests']
    if 'fillrate' not in cols and 'impressions' in cols:
        with np.errstate(divide='ignore', invalid='ignore'):
            df['fillrate'] = np.where(df['requests'] > 0, df['impressions'] / df['requests'], 0.0)

    return df


# --------------------------------------------------------------------------
# Section 4: Tag-level pipeline (also reused, tag-scoped, for bundles in Sec 6)
# --------------------------------------------------------------------------
def aggregate_by_group(df: pd.DataFrame, group_col: str, cfg: Config) -> pd.DataFrame:
    g = df.groupby(group_col, dropna=False)
    agg = g.agg(
        Requests=('requests', 'sum'),
        Revenue=('revenue', 'sum'),
        Impressions=('impressions', 'sum'),
        Date_Count=('Date', lambda s: s.nunique()),
    ).reset_index().rename(columns={group_col: 'Group'})

    agg['Date_Count'] = agg['Date_Count'].replace(0, 1)
    with np.errstate(divide='ignore', invalid='ignore'):
        agg['FillRate'] = np.where(agg['Requests'] > 0, agg['Impressions'] / agg['Requests'], 0.0)
    agg['Avg_Daily_Requests'] = agg['Requests'] / agg['Date_Count']
    agg['Avg_Daily_Revenue'] = agg['Revenue'] / agg['Date_Count']
    return agg


def derive_trend(df: pd.DataFrame, group_col: str, from_date, to_date) -> pd.Series:
    mid = from_date + (to_date - from_date) / 2
    half1 = df[df['Date'] <= mid].groupby(group_col)['revenue'].sum()
    half2 = df[df['Date'] > mid].groupby(group_col)['revenue'].sum()
    groups = pd.Index(df[group_col].unique())
    half1 = half1.reindex(groups, fill_value=0.0)
    half2 = half2.reindex(groups, fill_value=0.0)

    trend = pd.Series('Stable', index=groups)
    trend[half2 > half1 * 1.10] = 'Increasing'
    trend[half2 < half1 * 0.90] = 'Decreasing'
    return trend


def add_derived_and_scores(agg: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    agg = agg.copy()

    with np.errstate(divide='ignore', invalid='ignore'):
        agg['Rev_per_1k_req'] = np.where(agg['Requests'] > 0, 1000 * agg['Revenue'] / agg['Requests'], 0.0)
    agg['log_Requests'] = np.log1p(agg['Requests'])

    if cfg.cost_enabled:
        if cfg.cost_basis == 'requests':
            agg['Server_Cost'] = (agg['Requests'] / 1_000_000) * cfg.cost_per_1m_req
        else:
            agg['Server_Cost'] = (agg['Impressions'] / 1_000) * cfg.cost_per_1k_imp
        agg['Net_Revenue'] = agg['Revenue'] - agg['Server_Cost']
    else:
        agg['Server_Cost'] = 0.0
        agg['Net_Revenue'] = agg['Revenue']

    rev_p75 = agg['Avg_Daily_Revenue'].quantile(0.75)
    rev_p75 = rev_p75 if rev_p75 > 0 else 1e-9
    agg['Rev_Weight'] = (0.5 + 0.3 * np.minimum(agg['Avg_Daily_Revenue'] / rev_p75, 1.0)).clip(0.5, 0.8)
    agg['Fill_Weight'] = 1.0 - agg['Rev_Weight']

    fill_nonzero = agg.loc[agg['FillRate'] != 0, 'FillRate'].values
    rev_nonzero = agg.loc[agg['Rev_per_1k_req'] != 0, 'Rev_per_1k_req'].values
    median_fill = np.median(fill_nonzero) if len(fill_nonzero) else 0.0
    std_fill = robust_std(fill_nonzero) if len(fill_nonzero) else 1e-9
    median_rev = np.median(rev_nonzero) if len(rev_nonzero) else 0.0
    std_rev = robust_std(rev_nonzero) if len(rev_nonzero) else 1e-9
    median_log_req = np.median(agg['log_Requests'].values)
    std_log_req = robust_std(agg['log_Requests'].values)

    agg['FillRate_Perf_Score'] = np.where(
        agg['FillRate'] == 0, 0.0,
        10 * np.tanh((agg['FillRate'] - median_fill) / std_fill)
    )
    agg['Revenue_Perf_Score'] = np.where(
        agg['Rev_per_1k_req'] == 0, 0.0,
        10 * np.tanh((agg['Rev_per_1k_req'] - median_rev) / std_rev)
    )

    agg['Confidence_Weight'] = 0.5 * (1 + np.tanh((agg['log_Requests'] - median_log_req) / std_log_req))

    agg['Fill_Final_Score'] = agg['FillRate_Perf_Score'] * agg['Confidence_Weight']
    agg['Revenue_Final_Score'] = agg['Revenue_Perf_Score'] * agg['Confidence_Weight']
    agg['Final_Value_Score'] = agg['Fill_Weight'] * agg['Fill_Final_Score'] + agg['Rev_Weight'] * agg['Revenue_Final_Score']

    agg['Load_Pct'] = pct_rank(agg['Requests'])
    agg['Value_Pct'] = pct_rank(agg['Final_Value_Score'])

    return agg


# --------------------------------------------------------------------------
# Section 5: Decision layer
# --------------------------------------------------------------------------
def score_sign(score):
    if score < -0.5:
        return 'S-'
    if score > 0.5:
        return 'S+'
    return 'S0'


def magnitude_tier(abs_score_series: pd.Series) -> pd.Series:
    if abs_score_series.nunique() <= 1:
        return pd.Series('Med', index=abs_score_series.index)
    q1 = abs_score_series.quantile(1 / 3)
    q2 = abs_score_series.quantile(2 / 3)
    tier = pd.Series('Med', index=abs_score_series.index)
    tier[abs_score_series <= q1] = 'Low'
    tier[abs_score_series >= q2] = 'High'
    return tier


def load_tier(load_pct):
    if load_pct < 0.25:
        return 'L1'
    if load_pct < 0.75:
        return 'L2'
    return 'L3'


def confidence_tier(conf_series: pd.Series) -> pd.Series:
    q25 = conf_series.quantile(0.25)
    q75 = conf_series.quantile(0.75)
    tier = pd.Series('Normal', index=conf_series.index)
    tier[conf_series <= q25] = 'Low'
    tier[conf_series >= q75] = 'High'
    return tier


BASE_TABLE_SMINUS = {
    ('High', 'L1'): 0.50, ('High', 'L2'): 0.25, ('High', 'L3'): 0.00,
    ('Med', 'L1'): 0.75, ('Med', 'L2'): 0.50, ('Med', 'L3'): 0.25,
    ('Low', 'L1'): 1.00, ('Low', 'L2'): 0.75, ('Low', 'L3'): 0.50,
}
BASE_TABLE_S0 = {'L1': 1.00, 'L2': 1.00, 'L3': 0.75}
BASE_TABLE_SPLUS = {
    ('High', 'L1'): 1.75, ('High', 'L2'): 1.50, ('High', 'L3'): 1.25,
    ('Med', 'L1'): 1.50, ('Med', 'L2'): 1.25, ('Med', 'L3'): 1.00,
    ('Low', 'L1'): 1.00, ('Low', 'L2'): 1.00, ('Low', 'L3'): 1.00,
}


def ladder_step(m, direction):
    idx = LADDER.index(min(LADDER, key=lambda x: abs(x - m)))
    if direction == 'up':
        idx = min(idx + 1, len(LADDER) - 1)
    elif direction == 'down':
        idx = max(idx - 1, 0)
    return LADDER[idx]


def run_decision_layer(agg: pd.DataFrame, trend: pd.Series, cfg: Config, rev_p25_override=None) -> pd.DataFrame:
    agg = agg.copy()
    agg['AbsScore'] = agg['Final_Value_Score'].abs()
    agg['ScoreSign'] = agg['Final_Value_Score'].apply(score_sign)
    agg['Mag_Tier'] = magnitude_tier(agg['AbsScore'])
    agg['Load_Tier'] = agg['Load_Pct'].apply(load_tier)
    agg['Conf_Tier'] = confidence_tier(agg['Confidence_Weight'])
    agg['Trend'] = agg['Group'].map(trend).fillna('Stable')

    rev_p25 = rev_p25_override if rev_p25_override is not None else agg['Avg_Daily_Revenue'].quantile(0.25)

    mults, actions, reasons = [], [], []
    for _, row in agg.iterrows():
        sign = row['ScoreSign']
        if sign == 'S-':
            m = BASE_TABLE_SMINUS[(row['Mag_Tier'], row['Load_Tier'])]
        elif sign == 'S+':
            m = BASE_TABLE_SPLUS[(row['Mag_Tier'], row['Load_Tier'])]
        else:
            m = BASE_TABLE_S0[row['Load_Tier']]

        if row['Trend'] == 'Increasing':
            m = ladder_step(m, 'up')
        elif row['Trend'] == 'Decreasing' and sign == 'S-':
            m = ladder_step(m, 'down')

        action = 'Block' if m == 0.00 else ('Monitor' if m == 1.00 else ('Decrease QPS' if m < 1.00 else 'No Change (Monitor)'))

        if action == 'Block' and row['Revenue_Final_Score'] > 0:
            m, action = 0.25, 'Decrease QPS'
        if action == 'Block' and row['Conf_Tier'] == 'Low':
            m, action = 0.50, 'Decrease QPS'
        if action == 'Block' and row['Trend'] == 'Increasing':
            m, action = 0.25, 'Decrease QPS'
        if action == 'Block' and row['Avg_Daily_Revenue'] > rev_p25:
            m, action = 0.25, 'Decrease QPS'

        if row['Final_Value_Score'] > 0.5 and m < 1.00:
            m, action = 1.00, 'Monitor'

        if cfg.cost_enabled and row['Net_Revenue'] <= 0:
            m, action = 0.00, 'Block'

        reason = ''
        if cfg.rev_threshold_enabled and action == 'Block' and row['Revenue'] < cfg.rev_threshold:
            m, action = 1.00, 'Monitor'
            reason = 'Revenue threshold override'

        mults.append(m)
        actions.append(action)
        reasons.append(reason)

    agg['Final_Mult'] = mults
    agg['Suggested_Action'] = actions
    agg['Override_Reason'] = reasons
    agg['QPS_Change_Pct'] = np.where(agg['Suggested_Action'] == 'No Change (Monitor)', 0.0, (agg['Final_Mult'] - 1) * 100)
    agg['Daily_Cap'] = np.where(agg['Final_Mult'] < 1.00, agg['Avg_Daily_Requests'] * agg['Final_Mult'], np.nan)
    # Hook for a future capping API: replace Daily_Cap's implicit "current cap" assumption
    # (currently == Avg_Daily_Requests) with a real current-cap value pulled live, then
    # Suggested_Cap_Change = Daily_Cap - CURRENT_CAP to recommend increase/decrease deltas.

    return agg


def apply_revenue_loss_cap(agg: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    agg = agg.copy()
    agg['Projected_Rev_Loss'] = 0.0
    agg['Reverted_By_Cap'] = False

    is_block = agg['Suggested_Action'] == 'Block'
    is_decrease = agg['Suggested_Action'] == 'Decrease QPS'
    agg.loc[is_block, 'Projected_Rev_Loss'] = agg.loc[is_block, 'Revenue']
    agg.loc[is_decrease, 'Projected_Rev_Loss'] = agg.loc[is_decrease, 'Revenue'] * (1 - agg.loc[is_decrease, 'Final_Mult'])

    if not cfg.max_rev_loss_enabled:
        return agg

    total_revenue = agg['Revenue'].sum()
    rev_loss_budget = (cfg.max_rev_loss_pct / 100) * total_revenue
    total_projected_loss = agg['Projected_Rev_Loss'].sum()

    if total_projected_loss <= rev_loss_budget:
        return agg

    down = agg[is_block | is_decrease].sort_values(
        by=['Final_Value_Score', 'Projected_Rev_Loss'], ascending=[False, False]
    )

    remaining = total_projected_loss
    for idx in down.index:
        if remaining <= rev_loss_budget:
            break
        remaining -= agg.loc[idx, 'Projected_Rev_Loss']
        agg.loc[idx, 'Final_Mult'] = 1.00
        agg.loc[idx, 'Suggested_Action'] = 'Monitor'
        agg.loc[idx, 'Reverted_By_Cap'] = True
        agg.loc[idx, 'QPS_Change_Pct'] = 0.0
        agg.loc[idx, 'Daily_Cap'] = np.nan
        agg.loc[idx, 'Projected_Rev_Loss'] = 0.0

    return agg


# --------------------------------------------------------------------------
# Section 6: Bundle scoring / blocklist
# --------------------------------------------------------------------------
def score_bundles_for_tag(df_tag_rows: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    bundle_agg = aggregate_by_group(df_tag_rows, 'bundle', cfg)
    if bundle_agg.empty:
        return bundle_agg
    bundle_agg = add_derived_and_scores(bundle_agg, cfg)
    trend = pd.Series('Stable', index=bundle_agg['Group'])
    rev_p25_b = bundle_agg['Avg_Daily_Revenue'].quantile(0.25)
    bundle_agg = run_decision_layer(bundle_agg, trend, cfg, rev_p25_override=rev_p25_b)
    bundle_agg['bundle_rev_p25'] = rev_p25_b
    return bundle_agg


def bundle_blocklist_for_tag(df_tag_rows: pd.DataFrame, cfg: Config) -> str:
    bundle_scores = score_bundles_for_tag(df_tag_rows, cfg)
    if bundle_scores.empty:
        return ''
    qualifying = bundle_scores[
        (bundle_scores['Suggested_Action'] == 'Block')
        & (bundle_scores['Avg_Daily_Revenue'] <= bundle_scores['bundle_rev_p25'])
    ]
    return ', '.join(sorted(qualifying['Group'].astype(str).tolist()))


# --------------------------------------------------------------------------
# Full side pipeline (Sections 4-6 combined)
# --------------------------------------------------------------------------
def run_side(df_clean: pd.DataFrame, group_key: str, cfg: Config, has_bundle: bool) -> pd.DataFrame:
    date_mask = pd.Series(True, index=df_clean.index)
    if cfg.from_date is not None:
        date_mask &= df_clean['Date'] >= cfg.from_date
    if cfg.to_date is not None:
        date_mask &= df_clean['Date'] <= cfg.to_date
    df = df_clean[date_mask].copy()

    if df.empty or group_key not in df.columns:
        return pd.DataFrame()

    agg = aggregate_by_group(df, group_key, cfg)
    agg = add_derived_and_scores(agg, cfg)
    trend = derive_trend(df, group_key, cfg.from_date or df['Date'].min(), cfg.to_date or df['Date'].max())
    agg = run_decision_layer(agg, trend, cfg)
    agg = apply_revenue_loss_cap(agg, cfg)

    if has_bundle:
        blocklists = []
        for _, row in agg.iterrows():
            if row['Suggested_Action'] in ('Decrease QPS', 'Block'):
                tag_rows = df[df[group_key] == row['Group']]
                blocklists.append(bundle_blocklist_for_tag(tag_rows, cfg))
            else:
                blocklists.append('')
        agg['Bundle_Blocklist'] = blocklists
    else:
        agg['Bundle_Blocklist'] = ''

    agg.loc[agg['Suggested_Action'] == 'No Change (Monitor)', 'Bundle_Blocklist'] = ''
    agg.loc[agg['Suggested_Action'] == 'Monitor', 'Bundle_Blocklist'] = ''
    agg.loc[agg['Suggested_Action'] == 'Block', 'Bundle_Blocklist'] = ''

    return agg


# ==========================================================================
# Workbook builder (colour-coded Supply / Demand / Summary tabs)
# ==========================================================================
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(color='FFFFFF', bold=True, size=13)
SUBVALUE_FILL = PatternFill('solid', fgColor='D9EAF7')

BAND_FILL = PatternFill('solid', fgColor='EAF4FB')
NO_FILL = PatternFill(fill_type=None)

FILL_BLOCK = PatternFill('solid', fgColor='C00000')
FILL_DECREASE = PatternFill('solid', fgColor='FF7A00')
FILL_MONITOR = PatternFill('solid', fgColor='4472C4')
FILL_NOCHANGE = PatternFill('solid', fgColor='EDEDED')
FILL_BUNDLE_YELLOW = PatternFill('solid', fgColor='FFF2CC')
FILL_TOP_HEADER = PatternFill('solid', fgColor='375623')
FILL_BOTTOM_HEADER = PatternFill('solid', fgColor='E00000')

FONT_ACTION_WHITE = Font(color='FFFFFF', bold=True)
FONT_NOCHANGE = Font(color='404040', bold=True)
FONT_TREND_UP = Font(color='2E7D32', bold=True)
FONT_TREND_DOWN = Font(color='C00000', bold=True)

ACTION_FILL = {
    'Block': FILL_BLOCK,
    'Decrease QPS': FILL_DECREASE,
    'Monitor': FILL_MONITOR,
    'No Change (Monitor)': FILL_NOCHANGE,
}
ACTION_FONT = {
    'Block': FONT_ACTION_WHITE,
    'Decrease QPS': FONT_ACTION_WHITE,
    'Monitor': FONT_ACTION_WHITE,
    'No Change (Monitor)': FONT_NOCHANGE,
}

OUTPUT_COLUMNS = [
    ('Tag Name', 'Group'),
    ('Avg Daily Requests', 'Avg_Daily_Requests'),
    ('Avg Daily Revenue ($)', 'Avg_Daily_Revenue'),
    ('Trend', 'Trend'),
    ('Suggested Action', 'Suggested_Action'),
    ('QPS Change %', 'QPS_Change_Pct'),
    ('Daily Cap (Requests)', 'Daily_Cap'),
    ('Bundle Blocklist', 'Bundle_Blocklist'),
    ('Softened_By_Cap', 'Reverted_By_Cap'),
    ('Final Score', 'Final_Value_Score'),
    ('Load', 'Load_Tier'),
    ('Mag', 'Mag_Tier'),
]

COL_WIDTHS = [39, 20, 23, 12, 18, 14, 22, 42, 17, 13, 10, 10]

NUMFMT_REQ = '#,##0'
NUMFMT_USD = '$#,##0.00'
NUMFMT_PCT = '+0"%";-0"%";0"%"'
NUMFMT_SCORE = '0.00'


def _round_val(col, v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    if col == 'Final_Value_Score':
        return round(float(v), 2)
    if col in ('Avg_Daily_Requests', 'Daily_Cap'):
        return round(float(v), 0)
    if col == 'Avg_Daily_Revenue':
        return round(float(v), 2)
    return v


def write_side_sheet(wb, sheet_name, tag_col_label, agg: pd.DataFrame, cfg: Config):
    ws = wb.create_sheet(sheet_name)
    headers = [label if col != 'Group' else tag_col_label for label, col in OUTPUT_COLUMNS]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = 'A2'

    if not agg.empty:
        agg_sorted = agg.sort_values('Final_Value_Score')
        for i, (_, row) in enumerate(agg_sorted.iterrows()):
            values = []
            for label, col in OUTPUT_COLUMNS:
                v = row.get(col, '')
                v = _round_val(col, v)
                values.append(v if v is not None else '')
            ws.append(values)
            r = ws.max_row

            band = BAND_FILL if i % 2 == 0 else NO_FILL
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = band

            ws.cell(row=r, column=2).number_format = NUMFMT_REQ
            ws.cell(row=r, column=3).number_format = NUMFMT_USD
            ws.cell(row=r, column=6).number_format = NUMFMT_PCT
            ws.cell(row=r, column=7).number_format = NUMFMT_REQ
            ws.cell(row=r, column=10).number_format = NUMFMT_SCORE

            action = row['Suggested_Action']
            action_cell = ws.cell(row=r, column=5)
            action_cell.fill = ACTION_FILL.get(action, band)
            action_cell.font = ACTION_FONT.get(action, Font())

            trend_cell = ws.cell(row=r, column=4)
            if row['Trend'] == 'Increasing':
                trend_cell.font = FONT_TREND_UP
            elif row['Trend'] == 'Decreasing':
                trend_cell.font = FONT_TREND_DOWN

            blocklist_cell = ws.cell(row=r, column=8)
            if row.get('Bundle_Blocklist'):
                blocklist_cell.fill = FILL_BUNDLE_YELLOW

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_top_bottom_sheet(wb, sheet_name, tag_col_label, agg: pd.DataFrame):
    ws = wb.create_sheet(sheet_name)
    cols = OUTPUT_COLUMNS[:7]
    headers = [label if col != 'Group' else tag_col_label for label, col in cols]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    def write_section(title, fill, rows):
        ws.append([title] + [''] * (len(headers) - 1))
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).font = HEADER_FONT
        for i, (_, row) in enumerate(rows.iterrows()):
            values = []
            for label, col in cols:
                v = row.get(col, '')
                v = _round_val(col, v)
                values.append(v if v is not None else '')
            ws.append(values)
            rr = ws.max_row
            band = NO_FILL if i % 2 == 0 else BAND_FILL
            for c in range(1, len(headers) + 1):
                ws.cell(row=rr, column=c).fill = band
            ws.cell(row=rr, column=2).number_format = NUMFMT_REQ
            ws.cell(row=rr, column=3).number_format = NUMFMT_USD
            ws.cell(row=rr, column=6).number_format = NUMFMT_PCT
            ws.cell(row=rr, column=7).number_format = NUMFMT_REQ

            action = row['Suggested_Action']
            action_cell = ws.cell(row=rr, column=5)
            action_cell.fill = ACTION_FILL.get(action, band)
            action_cell.font = ACTION_FONT.get(action, Font())

    if not agg.empty:
        keepable = agg[~agg['Suggested_Action'].isin(['Block', 'Decrease QPS'])]
        top10 = keepable.sort_values('Final_Value_Score', ascending=False).head(10)
        bottom10 = agg.sort_values('Final_Value_Score', ascending=True).head(10)
        side_label = tag_col_label.replace(' Name', '')
        write_section(f'TOP 10 - Best Performing {side_label}s', FILL_TOP_HEADER, top10)
        ws.append([''] * len(headers))
        write_section(f'BOTTOM 10 - Worst Performing {side_label}s', FILL_BOTTOM_HEADER, bottom10)

    widths = [42, 20, 23, 12, 18, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(wb, cfg: Config, sides: list, from_date, to_date):
    ws = wb.create_sheet('Summary')
    n_cols = 1 + len(sides)

    def side_stats(agg: pd.DataFrame):
        if agg is None or agg.empty:
            return dict(n=0, rev=0.0, req=0.0, rev_impact=0.0, req_impact=0.0,
                        blocked=0, decreased=0, monitored=0, tags_with_blocklist=0,
                        total_bundles=0)
        total_rev = agg['Revenue'].sum() if 'Revenue' in agg.columns else 0.0
        total_req = agg['Requests'].sum() if 'Requests' in agg.columns else 0.0
        dec_mask = agg['Suggested_Action'].isin(['Decrease QPS', 'Block'])
        rev_impact = agg.loc[dec_mask, 'Projected_Rev_Loss'].sum()
        req_reduced = agg.loc[dec_mask, 'Requests'] - agg.loc[dec_mask, 'Requests'] * agg.loc[dec_mask, 'Final_Mult']
        req_impact = req_reduced.sum()
        bundle_counts = agg['Bundle_Blocklist'].apply(lambda s: len([x for x in str(s).split(',') if x.strip()]))
        monitored = agg['Suggested_Action'].isin(['Monitor', 'No Change (Monitor)']).sum()
        return dict(
            n=len(agg),
            rev=total_rev,
            req=total_req,
            rev_impact=(rev_impact / total_rev * 100) if total_rev else 0.0,
            req_impact=(req_impact / total_req * 100) if total_req else 0.0,
            blocked=(agg['Suggested_Action'] == 'Block').sum(),
            decreased=(agg['Suggested_Action'] == 'Decrease QPS').sum(),
            monitored=monitored,
            tags_with_blocklist=(bundle_counts > 0).sum(),
            total_bundles=bundle_counts.sum(),
        )

    stats = [side_stats(agg) for _, agg in sides]
    days = (to_date - from_date).days + 1
    side_names = ' + '.join(label.replace(' Tags', '') for label, _ in sides) or 'Supply + Demand'

    title = (f'Inventory Scoring - {side_names} | '
             f'{from_date.strftime("%d %b %Y")} | {to_date.strftime("%d %b %Y")} | V4 Bundle Logic')
    ws.append([title] + [''] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal='center')

    NUMFMT_PCT_PLAIN = '0.00"%"'

    def kv_row(label, values, number_format=None):
        ws.append([label] + list(values))
        r = ws.max_row
        ws.cell(row=r, column=1).fill = HEADER_FILL
        ws.cell(row=r, column=1).font = HEADER_FONT
        for i in range(len(values)):
            cell = ws.cell(row=r, column=2 + i)
            cell.fill = SUBVALUE_FILL
            if number_format and values[i] is not None and not isinstance(values[i], str):
                cell.number_format = number_format

    date_str = f'{from_date.strftime("%d %b %Y")} - {to_date.strftime("%d %b %Y")}'
    kv_row('Date Range', [date_str] * len(sides))
    kv_row('Total Tags', [s['n'] for s in stats])
    kv_row(f'Total Requests ({days}-day)', [round(s['req'], 0) for s in stats], NUMFMT_REQ)
    kv_row(f'Total Revenue ({days}-day)', [round(s['rev'], 2) for s in stats], NUMFMT_USD)
    kv_row('Revenue Impact % - Decrease/Block', [round(s['rev_impact'], 2) for s in stats], NUMFMT_PCT_PLAIN)
    kv_row('Requests Impact % - Decrease/Block', [round(s['req_impact'], 2) for s in stats], NUMFMT_PCT_PLAIN)
    kv_row('Revenue Loss Cap %', [cfg.max_rev_loss_pct if cfg.max_rev_loss_enabled else None for _ in stats], NUMFMT_PCT_PLAIN)
    kv_row('Tags with Bundle Blocklist', [s['tags_with_blocklist'] for s in stats])
    kv_row('Total Bundles to Block', [s['total_bundles'] for s in stats])

    ws.append([''] * n_cols)

    header_row = ['Suggested Action'] + [label for label, _ in sides]
    ws.append(header_row)
    r = ws.max_row
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    kv_row('Monitor', [s['monitored'] for s in stats])
    kv_row('Decrease QPS', [s['decreased'] for s in stats])
    kv_row('Block', [s['blocked'] for s in stats])

    ws.append([''] * n_cols)

    ws.append(['Bundle Logic Rules'] + [''] * (n_cols - 1))
    r = ws.max_row
    for c in range(1, n_cols + 1):
        ws.cell(row=r, column=c).fill = HEADER_FILL
        ws.cell(row=r, column=c).font = HEADER_FONT

    rules = [
        ('Bundle scoring scope', 'Only run for tags flagged Decrease QPS or Block'),
        ('Block condition', "Bundle Suggested Action = Block AND Avg Daily Revenue <= tag's bundle P25 revenue"),
        ('Revenue protection', 'Bundles earning above P25 daily revenue are never blocked'),
        ('No Change (Monitor)', 'Score favors a cap increase, but with no live capping API there is no '
                                 'current-cap baseline to compute a real increase from -- treated as Monitor, '
                                 'manual cap increase may be worth considering'),
        ('Yellow row', 'Tag has a bundle blocklist -- see Bundle Blocklist column'),
    ]
    for label, text in rules:
        row = [label, text] + [''] * (n_cols - 2)
        ws.append(row)

    ws.freeze_panes = 'A3'
    ws.column_dimensions['A'].width = 42
    for i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def build_workbook(raw: pd.DataFrame, cfg: Config):
    """End-to-end: raw DataFrame + Config -> openpyxl Workbook + warnings list.

    This is the single entry point used by both the CLI (run_agent.py) and
    the browser app (docs/assets/pipeline.py, executed via Pyodide).
    """
    cols = detect_columns(raw)
    warnings = []
    if 'supply_tag' not in cols:
        warnings.append("Supply Tag column not found -- Supply-side run skipped.")
    if 'demand_tag' not in cols:
        warnings.append("Demand Tag column not found -- Demand-side run skipped.")
    if 'bundle' not in cols:
        warnings.append("App Bundle column not found -- bundle scoring skipped.")
    if 'date' not in cols:
        raise ValueError("FATAL: Date column not found. Cannot proceed.")
    if 'requests' not in cols:
        raise ValueError("FATAL: Requests column not found. Cannot proceed.")
    if 'revenue' not in cols:
        raise ValueError("FATAL: Revenue column not found. Cannot proceed.")
    if 'fillrate' not in cols and 'impressions' not in cols:
        raise ValueError("FATAL: Neither FillRate nor Impressions column found. Cannot proceed.")

    df_clean = clean_rows(raw, cols)
    has_bundle = 'bundle' in cols
    if has_bundle:
        df_clean['bundle'] = raw[cols['bundle']].fillna('Unknown').astype(str)
    else:
        df_clean['bundle'] = ''

    if 'supply_tag' in cols:
        df_clean['supply_tag'] = raw[cols['supply_tag']].astype(str)
    if 'demand_tag' in cols:
        df_clean['demand_tag'] = raw[cols['demand_tag']].astype(str)

    from_date = cfg.from_date or df_clean['Date'].min()
    to_date = cfg.to_date or df_clean['Date'].max()
    cfg.from_date, cfg.to_date = from_date, to_date

    supply_agg = run_side(df_clean, 'supply_tag', cfg, has_bundle) if 'supply_tag' in cols else pd.DataFrame()
    demand_agg = run_side(df_clean, 'demand_tag', cfg, has_bundle) if 'demand_tag' in cols else pd.DataFrame()

    wb = Workbook()
    wb.remove(wb.active)

    sides_for_summary = []
    if not supply_agg.empty:
        write_side_sheet(wb, 'Supply Tag Analysis', 'Supply Tag Name', supply_agg, cfg)
        write_top_bottom_sheet(wb, 'Supply Top & Bottom 10', 'Supply Tag Name', supply_agg)
        sides_for_summary.append(('Supply Tags', supply_agg))
    if not demand_agg.empty:
        write_side_sheet(wb, 'Demand Tag Analysis', 'Demand Tag Name', demand_agg, cfg)
        write_top_bottom_sheet(wb, 'Top & Bottom 10', 'Demand Tag Name', demand_agg)
        sides_for_summary.append(('Demand Tags', demand_agg))

    write_summary_sheet(wb, cfg, sides_for_summary, from_date, to_date)
    wb.move_sheet('Summary', offset=-(len(wb.sheetnames) - 1))

    return wb, warnings, {'supply': supply_agg, 'demand': demand_agg}
